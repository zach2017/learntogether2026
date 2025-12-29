import logging
from typing import Optional
import PyPDF2
from docx import Document as DocxDocument
import openpyxl

logger = logging.getLogger(__name__)


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF file"""
    try:
        text = []
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text.append(page.extract_text())
        return '\n'.join(text)
    except Exception as e:
        logger.error(f"Error extracting text from PDF: {e}")
        return ""


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from DOCX file"""
    try:
        doc = DocxDocument(file_path)
        text = []
        for para in doc.paragraphs:
            if para.text.strip():
                text.append(para.text)
        return '\n'.join(text)
    except Exception as e:
        logger.error(f"Error extracting text from DOCX: {e}")
        return ""


def extract_text_from_xlsx(file_path: str) -> str:
    """Extract text from XLSX file"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        text = []
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            text.append(f"Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                row_text = ' | '.join(str(cell) for cell in row if cell)
                if row_text.strip():
                    text.append(row_text)
        return '\n'.join(text)
    except Exception as e:
        logger.error(f"Error extracting text from XLSX: {e}")
        return ""


def extract_text_from_file(file_path: str, file_type: str) -> str:
    """Extract text from various file types"""
    file_type = file_type.lower()
    
    if 'pdf' in file_type:
        return extract_text_from_pdf(file_path)
    elif 'docx' in file_type or 'word' in file_type:
        return extract_text_from_docx(file_path)
    elif 'xlsx' in file_type or 'excel' in file_type or 'spreadsheet' in file_type:
        return extract_text_from_xlsx(file_path)
    elif 'text' in file_type or 'plain' in file_type:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            logger.error(f"Error reading text file: {e}")
            return ""
    else:
        logger.warning(f"Unsupported file type: {file_type}")
        return ""


def generate_summary(text: str, max_length: int = 300) -> str:
    """Generate a simple summary by extracting key sentences"""
    if not text or len(text.strip()) == 0:
        return "Unable to generate summary from document."
    
    # Simple extractive summary: take first N characters
    text = text.strip()
    
    if len(text) > max_length:
        summary = text[:max_length] + "..."
    else:
        summary = text
    
    return summary
